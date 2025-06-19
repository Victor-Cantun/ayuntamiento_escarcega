from django.shortcuts import render
from apps.a_users.models import Profile
# Create your views here.

from rest_framework.views import APIView

from rest_framework import status, generics, permissions
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework_simplejwt.views import TokenObtainPairView
from django.contrib.auth import get_user_model
#upload
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser, FormParser
from django.shortcuts import get_object_or_404
from .models import Document

from .serializers import UserRegistrationSerializer, UserLoginSerializer, UserSerializer, DocumentSerializer
from django.contrib.auth.decorators import login_required
#admin
from django.shortcuts import get_object_or_404
#list
from django.db.models import Count
#export-excel
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.utils import timezone

User = get_user_model()

class UserRegistrationView(generics.CreateAPIView):
    queryset = User.objects.all()
    serializer_class = UserRegistrationSerializer
    permission_classes = [AllowAny]

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        no_user = Profile.objects.filter(role=4).count()
        if no_user == 500:
            return Response({'error': 'Se alcanzó el limite de aspirantes, para mayor información acude al departamento de Recursos Humanos del H. Ayuntamiento de Escárcega con los documentos solicitados'}, status=status.HTTP_400_BAD_REQUEST)
        else:
            user = serializer.save()
            # Generar tokens JWT
            refresh = RefreshToken.for_user(user)
            
            return Response({
                'message': 'Usuario registrado exitosamente',
                'user': UserSerializer(user).data,
                'tokens': {
                    'refresh': str(refresh),
                    'access': str(refresh.access_token),
                }
            }, status=status.HTTP_201_CREATED)

class UserLoginView(generics.GenericAPIView):
    serializer_class = UserLoginSerializer
    permission_classes = [AllowAny]

    def post(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = serializer.validated_data['user']
        
        # Generar tokens JWT
        refresh = RefreshToken.for_user(user)
        
        return Response({
            'message': 'Inicio de sesión exitoso',
            'user': UserSerializer(user).data,
            'tokens': {
                'refresh': str(refresh),
                'access': str(refresh.access_token),
            }
        })

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def logout_view(request):
    try:
        refresh_token = request.data.get('refresh_token')
        if refresh_token:
            token = RefreshToken(refresh_token)
            token.blacklist()
        return Response({'message': 'Sesión cerrada exitosamente'})
    except Exception as e:
        return Response({'error': 'Token inválido'}, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def user_profile_view(request):
    serializer = UserSerializer(request.user)
    return Response(serializer.data)


""" class UploadPDFView(APIView):
    permission_classes = [AllowAny]
    def post(self, request, *args, **kwargs):

        serializer = DocumentSerializer(data=request.data)
        if serializer.is_valid():
            post = serializer.save()

        return Response({'message': 'Archivo subido correctamente'}, status=status.HTTP_200_OK) """

class DocumentViewSet(viewsets.ModelViewSet):
    serializer_class = DocumentSerializer
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]
    
    def get_queryset(self):
        """Solo mostrar documentos del usuario autenticado"""
        return Document.objects.filter(user=self.request.user)
    
    def perform_create(self, serializer):
        """Asignar el usuario actual al documento"""
        serializer.save(user=self.request.user)
    
    def create(self, request, *args, **kwargs):
        """Crear o actualizar documento por tipo"""
        print("Datos recibidos:", request.data)  # Debug
        print("Archivos recibidos:", request.FILES)  # Debug
        
        document_type = request.data.get('type')
        document_file = request.FILES.get('document')
        
        # Validaciones básicas
        if not document_type:
            return Response(
                {'error': 'El tipo de documento es requerido'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if not document_file:
            return Response(
                {'error': 'El archivo es requerido'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            document_type = int(document_type)
            if document_type not in [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11]:
                return Response(
                    {'error': 'Tipo de documento inválido. Debe ser 1, 2 o 3'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
        except (ValueError, TypeError):
            return Response(
                {'error': 'Tipo de documento debe ser un número válido'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Verificar si ya existe un documento de este tipo para el usuario
        existing_doc = Document.objects.filter(
            user=request.user, 
            type=document_type
        ).first()
        
        if existing_doc:
            # Actualizar documento existente
            try:
                serializer = self.get_serializer(existing_doc, data=request.data, partial=True)
                if serializer.is_valid():
                    serializer.save()
                    return Response({
                        'message': 'Documento actualizado exitosamente',
                        'document': serializer.data
                    }, status=status.HTTP_200_OK)
                else:
                    print("Errores de validación (actualización):", serializer.errors)  # Debug
                    return Response({
                        'error': 'Error de validación',
                        'details': serializer.errors
                    }, status=status.HTTP_400_BAD_REQUEST)
            except Exception as e:
                print("Error en actualización:", str(e))  # Debug
                return Response({
                    'error': f'Error interno al actualizar: {str(e)}'
                }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        else:
            # Crear nuevo documento
            try:
                serializer = self.get_serializer(data=request.data)
                if serializer.is_valid():
                    serializer.save(user=request.user)
                    return Response({
                        'message': 'Documento subido exitosamente',
                        'document': serializer.data
                    }, status=status.HTTP_201_CREATED)
                else:
                    print("Errores de validación (creación):", serializer.errors)  # Debug
                    return Response({
                        'error': 'Error de validación',
                        'details': serializer.errors
                    }, status=status.HTTP_400_BAD_REQUEST)
            except Exception as e:
                print("Error en creación:", str(e))  # Debug
                return Response({
                    'error': f'Error interno al crear: {str(e)}'
                }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['get'])
    def user_documents(self, request):
        """Obtener todos los documentos del usuario organizados por tipo"""
        documents = self.get_queryset()
        serializer = self.get_serializer(documents, many=True)
        
        # Organizar por tipo
        docs_by_type = {1: None, 2: None, 3: None, 4:None, 5:None, 6:None, 7:None, 8:None,9:None,10:None,11:None}
        for doc in serializer.data:
            docs_by_type[doc['type']] = doc
        
        return Response({
            'documents': docs_by_type,
            'total': documents.count()
        })
    
    def destroy(self, request, *args, **kwargs):
        """Eliminar documento"""
        instance = self.get_object()
        instance.delete()
        return Response({
            'message': 'Documento eliminado exitosamente'
        }, status=status.HTTP_200_OK)

@login_required
def police_admin(request):
    return render(request, "admin/police/index.html")
def police_applicants_list(request):
    applicants = User.objects.select_related("profile").filter(profile__role=4).annotate(total_documents=Count('documents')).order_by('id')
    return render(request, "admin/police/applicant/list.html",{"applicants":applicants})
def police_applicant_detail(request,pk):
    applicant = get_object_or_404(User, id=pk)
    documentos_existentes = Document.objects.filter(user_id = applicant.id)
   #tipos_documentos_map = {item[0]: item[1] for item in Document.types}
    #print(tipos_documentos_map)
    documentos_por_tipo = {doc.type: doc for doc in documentos_existentes}
    documentos_estado = []
    for doc_id, doc_name in Document.types:
        documento_encontrado = documentos_por_tipo.get(doc_id) 
        estado_documento = {
            'nombre': doc_name, # El nombre como 'acta', 'identificacion'
            'id': doc_id,       # El ID numérico (1, 2, 3)
            'existe': documento_encontrado is not None, # True si existe el documento
            'documento_info': None # Por defecto, no hay información
        }
        if documento_encontrado:
            # Si el documento existe, agregamos su información al diccionario
            estado_documento['documento_info'] = {
                'original_name': documento_encontrado.original_name,
                'document_url': documento_encontrado.document.url, # URL del archivo
                'maneja':documento_encontrado.utility_type,
                # Puedes añadir más campos aquí si los necesitas, por ejemplo:
                # 'id_db': documento_encontrado.id,
                # 'fecha_subida': documento_encontrado.upload_date, (si tuvieras este campo)
            }
        documentos_estado.append(estado_documento)
    context = {
        'documentos_estado': documentos_estado,
        'applicant': applicant,
    }
    return render(request, "admin/police/applicant/detail.html",context)

def export_list_employees(request):
    applicants = User.objects.select_related("profile").filter(profile__role=4).annotate(total_documents=Count('documents')).order_by('id')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Empleados"
        # Definir headers
    headers = [
        '#', 'Nombre', 'Apellido', 'correo electrónico', 'teléfono', 'No. de documentos'
    ]
        
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
        
    # Escribir headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        
    # Escribir datos
    count = 1
    for row, applicant in enumerate(applicants, 2):
        ws.cell(row=row, column=1, value=count)
        ws.cell(row=row, column=2, value=applicant.first_name)
        ws.cell(row=row, column=3, value=applicant.last_name)
        ws.cell(row=row, column=4, value=applicant.email)
        ws.cell(row=row, column=5, value=applicant.profile.phone)
        ws.cell(row=row, column=6, value=applicant.total_documents)
        count = count + 1
        #ws.cell(row=row, column=7, value=float(empleado.salario))
        #ws.cell(row=row, column=8, value=empleado.fecha_ingreso)
        #ws.cell(row=row, column=9, value="Sí" if empleado.activo else "No")
        
    # Ajustar ancho de columnas
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].auto_size = True
        ws.column_dimensions[column_letter].width = 15
        
    # Crear respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="aspirantes_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    # Guardar workbook en response
    wb.save(response)
    return response

